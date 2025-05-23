import openpyxl

arquivo = openpyxl.load_workbook('vendas.xlsx')
planilha = arquivo.active

total = 0
for linha in range(2, planilha.max_row + 1):
    valor = planilha.cell(row=linha, column=2).value
    try:
        total += float(valor)
    except (TypeError, ValueError):
        print(f"Valor inválido na linha {linha}: {valor}")

resumo = openpyxl.Workbook()
aba = resumo.active
aba.title = 'Resumo'
aba['A1'] = 'Total de Vendas'
aba['A2'] = total

resumo.save('resumo_vendas.xlsx')
print("Resumo criado com sucesso.")
